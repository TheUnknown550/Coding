from openpyxl import Workbook, load_workbook

wb = load_workbook('Python/3.9/Excel/Excels/NewExcel.xlsx')
ws = wb.active
ws['A1'].value = "Test"

wb.save('Python/3.9/Excel/Excels/NewExcel.xlsx')