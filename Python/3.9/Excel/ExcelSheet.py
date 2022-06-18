from openpyxl import load_workbook

wb = load_workbook('Python/3.9/Excel/Excels/NewExcel.xlsx')
ws = wb.active

ws1 = wb.create_sheet("Sheet_A")
ws1.title = "Title_A"

ws2 = wb["Title_A"]

ws2['A1'].value = 'Test'

wb.save('Python/3.9/Excel/Excels/NewExcel.xlsx')