from openpyxl import Workbook, load_workbook

wb = load_workbook('C:/Users/mattc/OneDrive/Documents/PRC/StudentPairs.xlsx')
ws = wb.active
ws['A2'].value = "Test"

wb.save('C:/Users/mattc/OneDrive/Documents/PRC/StudentPairs.xlsx')